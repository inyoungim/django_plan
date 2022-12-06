from datetime import datetime
from django.forms import model_to_dict
from django.http import JsonResponse, FileResponse
from rest_framework.decorators import api_view
from rest_framework import serializers
from django.db.models import Prefetch
from ..models import *
import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
from django.core.files.storage import FileSystemStorage
import shutil


class modelSpecSerializer(serializers.ModelSerializer):
    class Meta:
        model = modelSpec
        fields = '__all__'


class modelBasicSerializer(serializers.ModelSerializer):
    class Meta:
        model = modelBasic
        fields = '__all__'


class regionMasterSerializer(serializers.ModelSerializer):
    class Meta:
        model = regionMaster
        fields = '__all__'


def getDataOfList(item_list, column_name):
    for item in item_list:
        if item.column_name == column_name:
            return item

    return None

@api_view(['GET', 'POST', 'PUT', 'DELETE'])
def royaltySearch(request,site , pk):
    if request.method == "GET":
        active_page = int(request.GET.get('active_page', 1))
        page_num = int(request.GET.get('page_num', 10))
        print("page_num:", page_num)
        period_name = request.GET.get('period_name', '202210')
        basicSet = modelBasic.objects.using(site).filter(period_name=period_name)
        total_article = basicSet.count()  # 총 갯수 확인 할 것
        print("total_article:", total_article)
        total_page = (total_article // page_num) + 1
        if total_article % page_num == 0:
            total_page -= 1
        page_block = active_page // 5
        if active_page % 5 == 0: page_block -= 1
        previous_page = page_block * 5
        next_page = previous_page + 6
        if next_page > total_page:
            next_page = total_page
        if active_page > next_page:
            previous_page = 0
            active_page = 1

        page_list = []
        page_s = previous_page + 1
        while page_s <= next_page:
            page_list.append(page_s)
            page_s += 1
            if page_s > previous_page + 5: break
        if next_page in page_list:
            next_page = 0

        start_num = (active_page - 1) * page_num
        if start_num < 0:
            start_num = 0

        basicSet = basicSet.order_by('model_suffix')[start_num: start_num + page_num]
        basic_list = []
        for item in basicSet:
            basic_data = modelBasicSerializer(item, many=False, context={'site': site}).data
            option_data = modelOption.objects.using(site).filter(parent_id=basic_data["id"], column_data__startswith="Y")
            basic_data["royalty_option"] = option_data.count()
            basic_list.append(basic_data)

        data = {
            "data_list": basic_list,
            "page_list": page_list,
            "previous_page": previous_page,
            "active_page": active_page,
            "next_page": next_page,
            "page_num": page_num,
            "total_article": total_article,
        }
        return JsonResponse(data, safe=False)
    return JsonResponse(data={}, safe=False)


@api_view(['GET', 'POST', 'PUT', 'DELETE'])
def columnItem(request, site, pk):
    if request.method == "GET":
        db_data= modelSpec.objects.using(site).all()
        serializer = modelSpecSerializer(db_data, many=True, context={'site': site})
        data = {
            "data_list": serializer.data,
        }
        return JsonResponse(data, safe=False)
    elif request.method == "POST":
        column_name = request.data.get('column_name', '')
        column_data = request.data.get('column_data', '')
        print(column_name)
        try:
            nextid = modelSpec.objects.using(site).latest('id').id + 1
            nextno = modelSpec.objects.using(site).latest('order_no').order_no + 1
        except:
            nextid = 1
            nextno = 1

        db_data = modelSpec(
            id=nextid, order_no=nextno, column_name=column_name, column_data=column_data
        )
        print("site", site)
        db_data.save(using=site)
        request.preset['acc_history'] = {
            'acc_type1': 'CREATE',
            'acc_type2': 'salespo',
            'acc_type3': 'saveOrder',
            'acc_type4': f'{model_to_dict(db_data)}'
        }

        db_data = modelSpec.objects.using(site).all()
        serializer = modelSpecSerializer(db_data, many=True, context={'site': site})
        data = {
            "data_list": serializer.data,
        }

        return JsonResponse(data, safe=False)
    return JsonResponse(data={}, sefe=False)


@api_view(['GET', 'POST', 'PUT', 'DELETE'])
def uploadExcel(request, site, pk):
    if request.method == "POST":
        # 엑셀 업로드 (초기)
        files = request.data.get('attache_files', [])
        file_names = request.data.get('attache_file_names', [])
        input_mode = request.data.get('input_mode', '')
        period_name = request.data.get('period_name', '')
        if len(files) > 0:
            src = '/WEB_ROOT/DATA/temp/'
            des = f'/WEB_ROOT/{site}/royalty/itemInput/'
            os.makedirs(f'{des}', exist_ok=True)
            now = datetime.now()
            stime = now.strftime('%Y%m%d%H%M%S')
            region1_map = {}
            region1_master = regionMaster.objects.using(site).filter(mode='region1')
            for temp1 in region1_master:
                key = temp1.country
                val = temp1.region1
                region1_map[key] = val

            region2_map = {}
            region2_master = regionMaster.objects.using(site).filter(mode='region2')
            for temp2 in region2_master:
                key = temp2.code
                val = temp2.country
                region2_map[key] = val

            for i, file_name in enumerate(files):
                key_list = []
                key_list.append("Key")
                key_list.append("CSKD_보정")
                key_list.append("Sales Subsidiary")
                key_list.append("Model.Suffix")
                key_list.append("Production Model Suffix")
                key_list.append("Country")
                key_list.append("Region")
                key_list.append("Product Lvl.3")
                key_list.append("Product Lvl.4")
                key_list.append("Qty")
                db_data = modelSpec.objects.using(site).all()
                for data in db_data:
                    key_list.append(data.column_name)
                key_list.append("")
                key_list.append("")
                key_list.append("")
                key_list.append("")
                key_list.append("")
                key_list.append("")
                key_list.append("")



                extension = os.path.splitext(str(file_name))[1]
                filename = f'{stime}{extension}'
                excel_file = f'{des}{filename}'
                shutil.move(src + file_name, excel_file)
                wb = openpyxl.load_workbook(excel_file)
                for sheet_nm in wb.sheetnames:
                    sheet = wb[sheet_nm]
                    temp_data = []
                    bulk_create_list = []
                    # bulk_create_option_list = []
                    idk = 1
                    for row_data in sheet.iter_rows(min_row=2):
                        j = 0
                        try:
                            nextid = modelBasic.objects.using(site).latest('id').id + idk
                        except:
                            nextid = idk
                        new_item = modelBasic(id=nextid, period_name=period_name)
                        idk = idk + 1
                        for cell in row_data:
                            if j == 2:
                                new_item.sales_subsidiary = cell.value
                            elif j == 3:
                                new_item.model_suffix = cell.value
                            elif j == 4:
                                new_item.production_model_suffix = cell.value
                            elif j == 5:
                                new_item.country = cell.value
                            elif j == 6:
                                new_item.region = cell.value
                            elif j == 7:
                                new_item.product_lv3 = cell.value
                            elif j == 8:
                                new_item.product_lv4 = cell.value
                            elif j == 9:
                                new_item.qty = cell.value
                            elif j >= 10:
                                new_option_item = modelOption(parent_id=nextid)
                                new_option_item.column_name=key_list[j]
                                # bulk_create_option_list.append(bulk_create_option_list)
                                # print(bulk_create_option_list)
                            j = j + 1

                        temp_model = new_item.model_suffix
                        model_name = temp_model
                        temp_model = temp_model.replace("OLED", "")
                        new_item.inch = temp_model[0:2]
                        if temp_model.find("NANO") >= 0 or temp_model.find("QNED") >= 0:
                            new_item.series_name1 = temp_model[2:6]
                            new_item.series_name2 = temp_model[2:8]
                            new_item.nano_initial = temp_model[9:10]
                        elif model_name.find("OLED") >= 0:
                            new_item.series_name1 = temp_model[2:4]
                            new_item.series_name2 = temp_model[2:4]
                        else:
                            new_item.series_name1 = temp_model[2:4]
                            new_item.series_name2 = temp_model[2:6]

                        if temp_model.find("-") >= 0:
                            new_item.series_name3 = temp_model[2:temp_model.find("-")]
                            new_item.model_compr = model_name[:model_name.find("-")]
                            new_item.t2 = temp_model[temp_model.find("-")-1:temp_model.find("-")]
                            new_item.country_match = temp_model[temp_model.find("-")+1:temp_model.find(".")]
                        else:
                            new_item.series_name3 = temp_model[2:temp_model.find(".")]
                            new_item.model_compr = model_name[:model_name.find(".")]
                            new_item.t2 = temp_model[temp_model.find(".")-2:temp_model.find(".")-1]

                        new_item.suffix = model_name[model_name.find(".")+1:]
                        new_item.suffix_country = new_item.suffix[1:3]
                        try:
                            # print(new_item.suffix_country)
                            new_item.region2 = region2_map[new_item.suffix_country]
                        except:
                            pass
                        try:
                            new_item.region1 = region1_map[new_item.country]
                        except:
                            pass
                        if new_item.product_lv4.find("O/S") >= 0:
                            new_item.odm = "ODM"
                        else:
                            new_item.odm = "-"
                        if temp_model.find("OLED") >= 0:
                            new_item.product_family = "OLED"
                        elif temp_model[2:1] == "L":
                            new_item.product_family = "FHD"
                        else:
                            new_item.product_family = "UHD"
                        new_item.model_id = new_item.region2 + new_item.model_compr + new_item.region1
                        bulk_create_list.append(new_item)
                    modelBasic.objects.using(site).bulk_create(bulk_create_list)
    elif request.method == "PUT":
        files = request.data.get('attache_files', [])
        file_names = request.data.get('attache_file_names', [])
        input_mode = request.data.get('input_mode', '')
        period_name = request.data.get('period_name', '')
        if len(files) > 0:
            src = '/WEB_ROOT/DATA/temp/'
            des = f'/WEB_ROOT/{site}/royalty/itemResult/'
            os.makedirs(f'{des}', exist_ok=True)
            now = datetime.now()
            stime = now.strftime('%Y%m%d%H%M%S')
            for i, file_name in enumerate(files):
                key_list = []
                key_list.append("No")
                key_list.append("구분")
                key_list.append("법인")
                key_list.append("모델")
                key_list.append("-")
                key_list.append("-")
                key_list.append("-")
                key_list.append("-")
                key_list.append("모델")
                key_list.append("NANO QNED 연도")
                key_list.append("T2")
                key_list.append("-")
                key_list.append("-")
                key_list.append("국가")
                key_list.append("지역2")
                key_list.append("국가")
                key_list.append("지역1")
                key_list.append("Region")
                key_list.append("Product Lvl.4")
                key_list.append("ODM")
                key_list.append("구분")
                key_list.append("Qty")
                key_list.append("구분자")
                key_list.append("Main SoC_2")
                key_list.append("사")
                key_list.append("SW Service")
                key_list.append("Smart")
                key_list.append("3D")
                key_list.append("MR")
                key_list.append("Wifi BT")
                key_list.append("Analog")
                key_list.append("년도")
                db_data = modelSpec.objects.using(site).all()
                for data in db_data:
                    key_list.append(data.column_name)
                key_list.append("")
                key_list.append("")
                key_list.append("")
                key_list.append("")
                key_list.append("")
                key_list.append("")
                key_list.append("")

                extension = os.path.splitext(str(file_name))[1]
                filename = f'{stime}{extension}'
                excel_file = f'{des}{filename}'
                shutil.move(src + file_name, excel_file)
                wb = openpyxl.load_workbook(excel_file)
                bulk_insert_list = []
                bulk_update_list = []
                for sheet_nm in wb.sheetnames:
                    sheet = wb[sheet_nm]
                    id_list = []
                    for row_data in sheet.iter_rows(min_row=8):
                        j = 0
                        for cell in row_data:
                            if j == 0:
                                id_list.append(cell.value)
                                break
                            j = j + 1
                    print("id_list : ", id_list)
                    option_item_all = modelOption.objects.using(site).filter(
                        parent_id__in=id_list, column_name__in=key_list)
                    i = 0
                    for row_data in sheet.iter_rows(min_row=8):
                        j = 0
                        for cell in row_data:
                            if j >=31:
                                option_item = getDataOfList(option_item_all, key_list[j])
                                if option_item :
                                    option_item.column_data = cell.value
                                    bulk_update_list.append(option_item)
                                else:

                                    new_item = modelOption(parent_id=id_list[i], column_name=key_list[j],
                                                           column_data=cell.value)
                                    bulk_insert_list.append(new_item)
                            j = j + 1
                        i = i + 1

                modelOption.objects.using(site).bulk_create(bulk_insert_list,batch_size=500)
                modelOption.objects.using(site).bulk_update(bulk_update_list,['column_data'],batch_size=500)
    elif request.method == "GET":
        # file_download
        period_name = request.GET.get('period_name', '')
        basic_obj = modelBasic.objects.using(site).filter(period_name=period_name)
        # basic_obj = basic_obj.prefetch_related('modeloption_set')

        pre_basic_obj = modelBasic.objects.using(site).filter(period_name__lte=period_name)
        pre_basic_obj = pre_basic_obj.prefetch_related('modeloption_set')

        temp_head = []
        i = 1
        while i <= 35:
            temp_head.append("")
            i += 1

        head = []
        if request.preset["user_lang"] == "kr":
            head.append("No")
            head.append("구분")
            head.append("법인")
            head.append("모델")
            head.append("-")
            head.append("-")
            head.append("-")
            head.append("-")
            head.append("모델")
            head.append("NANO QNED 연도")
            head.append("T2")
            head.append("-")
            head.append("-")
            head.append("국가")
            head.append("지역2")
            head.append("국가")
            head.append("지역1")
            head.append("Region")
            head.append("Product Lvl.4")
            head.append("ODM")
            head.append("구분")
            head.append("Qty")
            head.append("구분자")
            head.append("Main SoC_2")
            head.append("사")
            head.append("SW Service")
            head.append("Smart")
            head.append("3D")
            head.append("MR")
            head.append("Wifi BT")
            head.append("Analog")
            head.append("년도")
        else:
            head.append("No")
            head.append("구분")
            head.append("법인")
            head.append("모델")
            head.append("-")
            head.append("-")
            head.append("-")
            head.append("-")
            head.append("모델")
            head.append("NANO QNED 연도")
            head.append("T2")
            head.append("-")
            head.append("-")
            head.append("국가")
            head.append("지역2")
            head.append("국가")
            head.append("지역1")
            head.append("Region")
            head.append("Product Lvl.4")
            head.append("ODM")
            head.append("구분")
            head.append("Qty")
            head.append("구분자")
            head.append("Main SoC_2")
            head.append("사")
            head.append("SW Service")
            head.append("Smart")
            head.append("3D")
            head.append("MR")
            head.append("Wifi BT")
            head.append("Analog")
            head.append("년도")

        db_data = modelSpec.objects.using(site).all()
        for data in db_data:
            head.append(data.column_name)
        white_font = Font(color="FFFFFF", size=9)
        black_font = Font(size=9, color="000000")
        black_color = PatternFill(start_color='000000', end_color='000000', patternType='solid')
        blue_color1 = PatternFill(start_color='DAEEF3', end_color='DAEEF3', patternType='solid')
        blue_color2 = PatternFill(start_color='B7DEE8', end_color='B7DEE8', patternType='solid')
        blue_color3 = PatternFill(start_color='92CDDC', end_color='92CDDC', patternType='solid')
        blue_color4 = PatternFill(start_color='31869B', end_color='31869B', patternType='solid')
        orange_color = PatternFill(start_color='E26B0A', end_color='E26B0A', patternType='solid')
        salgu_color = PatternFill(start_color='FDE9D9', end_color='FDE9D9', patternType='solid')
        green_color1 = PatternFill(start_color='00B050', end_color='00B050', patternType='solid')
        green_color2 = PatternFill(start_color='92D050', end_color='92D050', patternType='solid')
        yellow_color = PatternFill(start_color='FFFF00', end_color='FFFF00', patternType='solid')
        purple_color = PatternFill(start_color='C59EE2', end_color='C59EE2', patternType='solid')
        red_color = PatternFill(start_color='DA9694', end_color='DA9694', patternType='solid')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(temp_head) #1번 Row는 공백
        sheet.append(temp_head) #2번 Row는 공백
        sheet.append(temp_head) #3번 Row는 공백
        sheet.append(temp_head) #4번 Row는 공백
        sheet.append(temp_head) #5번 Row는 공백
        sheet.append(temp_head) #6번 Row는 공백
        sheet.append(head) #7번 Row 부터 타이틀 입력 ? 그냥 이렇게 해달라고 함.. 답답하네 ㅋㅋㅋ

        #여기도 개선 해야함..
        col_num = 0
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width
            col_num += 1

        sheet.cell(row=7, column=1).fill = black_color
        sheet.cell(row=7, column=1).font = white_font
        sheet.column_dimensions["A"].width = 4.38

        sheet.cell(row=7, column=2).font = black_font
        sheet.column_dimensions["B"].width = 7.38

        sheet.cell(row=7, column=3).font = black_font
        sheet.column_dimensions["C"].width = 6.13

        sheet.cell(row=7, column=4).font = black_font
        sheet.column_dimensions["D"].width = 18.13

        sheet.cell(row=7, column=5).fill = blue_color1
        sheet.cell(row=7, column=5).font = black_font
        sheet.column_dimensions["E"].width = 6.63

        sheet.cell(row=7, column=6).fill = blue_color2
        sheet.cell(row=7, column=6).font = black_font
        sheet.column_dimensions["F"].width = 6.5

        sheet.cell(row=7, column=7).fill = blue_color3
        sheet.cell(row=7, column=7).font = black_font
        sheet.column_dimensions["G"].width = 9.5

        sheet.cell(row=7, column=8).fill = blue_color4
        sheet.cell(row=7, column=8).font = black_font
        sheet.column_dimensions["H"].width = 13

        sheet.cell(row=7, column=9).fill = orange_color
        sheet.cell(row=7, column=9).font = black_font
        sheet.column_dimensions["I"].width = 13.13

        sheet.cell(row=7, column=10).fill = salgu_color
        sheet.cell(row=7, column=10).font = black_font
        sheet.column_dimensions["J"].width = 5.5

        sheet.cell(row=7, column=11).fill = black_color
        sheet.cell(row=7, column=11).font = white_font
        sheet.column_dimensions["K"].width = 8.5

        sheet.cell(row=7, column=12).fill = salgu_color
        sheet.column_dimensions["L"].width = 4.13

        sheet.cell(row=7, column=13).fill = salgu_color
        sheet.column_dimensions["M"].width = 8.25

        sheet.cell(row=7, column=14).fill = black_color
        sheet.cell(row=7, column=14).font = white_font
        sheet.column_dimensions["N"].width = 8.5

        sheet.cell(row=7, column=15).fill = green_color1
        sheet.column_dimensions["O"].width = 10

        sheet.column_dimensions["P"].width = 8.63

        sheet.cell(row=7, column=17).fill = green_color2
        sheet.column_dimensions["Q"].width = 8.63

        sheet.column_dimensions["R"].width = 6.75

        sheet.column_dimensions["S"].width = 13

        sheet.cell(row=7, column=20).fill = black_color
        sheet.cell(row=7, column=20).font = white_font
        sheet.column_dimensions["T"].width = 7

        sheet.cell(row=7, column=21).fill = black_color
        sheet.cell(row=7, column=21).font = white_font
        sheet.column_dimensions["U"].width = 7

        sheet.column_dimensions["V"].width = 7.38

        sheet.cell(row=7, column=23).fill = yellow_color
        sheet.column_dimensions["w"].width = 11

        sheet.cell(row=7, column=24).fill = purple_color
        sheet.cell(row=7, column=24).border = thin_border
        sheet.column_dimensions["X"].width = 6.38

        sheet.cell(row=7, column=25).fill = red_color
        sheet.cell(row=7, column=25).border = thin_border
        sheet.column_dimensions["Y"].width = 6.38

        sheet.cell(row=7, column=26).fill = red_color
        sheet.cell(row=7, column=26).border = thin_border
        sheet.column_dimensions["Z"].width = 6.38

        sheet.cell(row=7, column=27).fill = red_color
        sheet.cell(row=7, column=27).border = thin_border
        sheet.column_dimensions["AA"].width = 6.38

        sheet.cell(row=7, column=28).border = thin_border
        sheet.column_dimensions["AB"].width = 6.38

        sheet.cell(row=7, column=29).border = thin_border
        sheet.column_dimensions["AC"].width = 6.38

        sheet.cell(row=7, column=30).border = thin_border
        sheet.column_dimensions["AD"].width = 6.38

        sheet.cell(row=7, column=31).border = thin_border
        sheet.column_dimensions["AE"].width = 6.38

        sheet.cell(row=7, column=32).border = thin_border
        sheet.column_dimensions["AF"].width = 6.38

        # sheet["Sheet"].font = Font(size=9)

        row_num = 0
        pre_item_info = {} #Map으로 설정함
        for preitem in pre_basic_obj:
            gubun_id = preitem.region2
            gubun_id = gubun_id+preitem.model_compr
            gubun_id = gubun_id + preitem.region1
            pre_option = preitem.modeloption_set.all()
            for option in pre_option:
                temp_key = gubun_id+option.column_name
                if temp_key in pre_item_info:
                    pass
                else:
                    pre_item_info[temp_key] = option.column_data

        for item in basic_obj:
            row_num += 1
            col_data = []
            col_data.append(item.id)
            col_data.append(item.period_name)
            col_data.append(item.sales_subsidiary)
            col_data.append(item.model_suffix)
            col_data.append(item.inch)
            col_data.append(item.series_name1)
            col_data.append(item.series_name2)
            col_data.append(item.series_name3)
            col_data.append(item.model_compr)
            col_data.append(item.nano_initial)
            col_data.append(item.t2)
            col_data.append("")
            col_data.append(item.suffix)
            col_data.append(item.suffix_country)
            col_data.append(item.region2)
            col_data.append(item.country)
            col_data.append(item.region1)
            col_data.append(item.region)
            col_data.append(item.product_lv4)
            col_data.append(item.odm)
            col_data.append(item.product_family)
            col_data.append(item.qty)
            col_data.append(item.model_id)
            col_data.append(item.main_soc)
            col_data.append(item.soc_co)
            col_data.append(item.sw_service)
            col_data.append(item.smart)
            col_data.append(item.three_d)
            col_data.append(item.mr)
            col_data.append(item.wifi_bt)
            col_data.append(item.analog)
            col_data.append(item.year)
            # 여기를 개선 해야함..
            master_key = item.region2 + item.model_compr + item.region1
            for idx in range(len(head)):
                if idx > 35:
                    try:
                        key_val = pre_item_info[master_key+head[idx]]
                    except:
                        key_val = ""
                    col_data.append(key_val)

            sheet.append(col_data)

        new_filename = f'/WEB_ROOT/{site}/royalty/itemResult/{period_name}_royalty_item_list.xlsx'
        wb.save(new_filename)
        fs = FileSystemStorage(new_filename)
        response = FileResponse(fs.open(new_filename, 'rb'), content_type="application/octet-stream")
        response['Content-Disposition'] = f'attachment; filename={new_filename}'
        return response

    return JsonResponse(data={}, safe=False)

@api_view(['GET'])
def royaltyFormExcelDownload(request, site):
    file_path = f"/WEB_ROOT/DATA/{site}/royalty/royalty_form.xlsx"
    file_name = "royalty_form.xlsx"
    fs = FileSystemStorage(file_path)
    response = FileResponse(fs.open(file_path, 'rb'), content_type="application/octet-stream")
    response['Content-Disposition'] = f'attachment; filename={file_name}'
    return response